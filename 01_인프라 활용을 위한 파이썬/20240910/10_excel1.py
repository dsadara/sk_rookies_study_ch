from openpyxl import Workbook

wb = Workbook()

# 엑셀 시트 활성화
ws = wb.active
# 시트 이름 수정
ws.title = "SK 쉴더스 루키즈"

# 데이터 추가
ws['B2'] = '날짜'
ws.cell(row=2, column=1, value='제품명')

wb.save("01_엑셀자동화.xlsx")