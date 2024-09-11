from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime
import os

send_email = "fkelfkenl10@naver.com"
send_pwd = ""
recv_email = "fkelfkenl10@naver.com"

def mail_sender():
    smtp_name = "smtp.naver.com"
    smtp_port = 587
    now = datetime.now().strftime("%Y-%m-%d")

    # MIMEMultipart 객체 생성: 이메일 본문과 첨부 파일을 포함할 수 있음
    msg = MIMEMultipart()
    msg['Subject'] = f"{now} 수집된 보안 동향 정보"
    msg['From'] = send_email
    msg['To'] = recv_email

    # 이메일 본문 내용
    text = f"보안 동향 정보입니다. {now} 날짜에 수집된 정보를 기록하고 있습니다."
    email_body_part = MIMEText(text, 'plain', 'utf-8')
    msg.attach(email_body_part)

    # 첨부 파일 경로
    file_path = r"C:\python_ex\240911_excel.xlsx"

    with open(file_path, 'rb') as file:
        file_part = MIMEApplication(file.read())
        file_part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"' 
        msg.attach(file_part)

    email_string = msg.as_string()
    print(email_string)

    s = smtplib.SMTP(smtp_name, smtp_port)
    s.starttls()
    s.login(send_email, send_pwd)
    s.sendmail(send_email, recv_email, email_string)
    s.quit()

wb = Workbook()
ws = wb.active
ws['A1'] = "제목"
ws['B1'] = "URL 링크"

url = "https://www.malware-traffic-analysis.net/2024/index.html"
header_info = {'User-Agent': 'Mozilla/5.0'}

r = requests.get(url, headers=header_info)
soup = BeautifulSoup(r.text, 'lxml')

datas = soup.select("#main_content > div.blog_entry > ul > li > a.main_menu")

row_data = 2
for data in datas:
    title_text = data.get_text()
    title_url = f"https://www.malware-traffic-analysis.net/2024/{data.get('href')}"
    ws.cell(row=row_data, column=1, value=title_text)
    ws.cell(row=row_data, column=2, value=title_url)
    row_data += 1

wb.save("240911_excel.xlsx")
mail_sender()