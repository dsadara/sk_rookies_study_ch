"""
★문제 가이드★
1.   멀웨어 트래픽 분석 사이트에서 보안 관련 최신 정보를 스크래핑합니다.
https://www.malware-traffic-analysis.net/2024/index.html
2.   엑셀 파일의 첫 번째 열에는 제목, 두 번째 열에는 URL 링크가 포함되어야 합니다.
3.   스크래핑한 날짜를 제목에 포함하여 매일 해당 정보를 이메일로 전송합니다.
4.   매일 한 번씩 웹 스크래핑, 데이터 저장, 이메일 전송이 자동으로 실행되도록 스케줄링합니다.
"""
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import datetime
import time

# 분석 사이트 스크래핑
def scrap():
    url = "https://www.malware-traffic-analysis.net/2024"
    header_info = {'User-Agent': 'Mozilla/5.0'}

    r = requests.get(url, headers=header_info)
    soup = BeautifulSoup(r.text, 'lxml')
    datas = soup.select("#main_content > div.blog_entry > ul > li > a.main_menu")

    scrap_data = {}
    for data in datas:
        link_text = data.get_text()
        link_url = url + data.get('href')
        scrap_data[link_text] = link_url
    return scrap_data

# 엑셀 데이터 생성
def save_to_excel(scrap_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "scrap_result"

    for index, url in enumerate(scrap_data):
        ws.cell(row=index+1, column=1, value=url)
        ws.cell(row=index+1, column=2, value=scrap_data.get(url))

    wb.save("scrap_result.xlsx")

# 이메일 전송
def email_scrap_result():
    send_email = "fkelfkenl10@naver.com"
    send_pwd = "---"
    recv_email = "fkelfkenl10@naver.com"

    smtp_name = "smtp.naver.com"
    smtp_port = 587

    current_date = datetime.datetime.now().strftime("%Y-%m-%d")

    msg = MIMEMultipart()
    msg['Subject'] = f"[{current_date}]보안 정보 스크래핑 결과"
    msg['From'] = send_email
    msg['To'] = recv_email

    text = "멀웨어 트래픽 분석 사이트(https://www.malware-traffic-analysis.net/2024/index.html) 보안 관련 최신 정보"
    email_body_part = MIMEText(text, 'plain', 'utf-8')
    msg.attach(email_body_part)

    # 첨부 파일 경로
    file_path = r"C:\python_ex\scrap_result.xlsx"

    with open(file_path, 'rb') as file:
        file_part = MIMEApplication(file.read())
        file_part['Content-Disposition'] = 'attachment; filename="scrap_result.xlsx"' 
        msg.attach(file_part)

    email_string = msg.as_string()

    s = smtplib.SMTP(smtp_name, smtp_port)
    s.starttls()
    s.login(send_email, send_pwd)
    s.sendmail(send_email, recv_email, email_string)
    s.quit()

while True:
    scrap_result = scrap()
    save_to_excel(scrap_result)
    email_scrap_result()
    time.sleep(60)